#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import datetime, date, timedelta
from io import BytesIO
import os
import threading
import unicodedata
import logging
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InputFile
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)

BOT_TOKEN = "8334665305:AAFykh9AZ1d4wgmlze_b8kx5rk2XvgRRCCA"
ADMINS = {5318613615, 792085774}

EXCEL_FILE = "attendance.xlsx"
DATE_FORMAT = "%Y-%m-%d"
TIME_FORMAT = "%H:%M"

BTN_KELDIM = "✅ Keldim"
BTN_KETDIM = "🚪 Ketdim"
BTN_SABBLI = "⚠️ Sababli bo'ldi"

BTN_WEEK = "📅 Haftalik hisobot"
BTN_MONTH = "🗓 Oylik hisobot"
BTN_ASK = "⌨️ Sanani yozib so'rash"

WAITING_DATES = 1

excel_lock = threading.Lock()
TZ = ZoneInfo("Asia/Tashkent")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


def _normalize(text: str) -> str:
    if text is None:
        return ""
    t = unicodedata.normalize("NFKC", text)
    t = t.replace("\u2019", "'").replace("\u2018", "'").replace("\u02BC", "'")
    t = t.replace("\u201c", '"').replace("\u201d", '"')
    return t.strip()


def load_or_create():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws["A1"] = "Ism / Sana"
        ws["A2"] = ""
        try:
            ws.column_dimensions["A"].width = 12
        except Exception:
            pass
        wb.save(EXCEL_FILE)
        return wb
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    max_col = ws.max_column
    col = 2
    while col <= max_col - 1:
        try:
            h1 = ws.cell(row=1, column=col).value
            h2 = ws.cell(row=1, column=col + 1).value
        except Exception:
            h1 = h2 = None
        already = False
        for m in ws.merged_cells.ranges:
            try:
                if m.min_row == 1 and m.max_row == 1 and m.min_col <= col and m.max_col >= col + 1:
                    already = True
                    break
            except Exception:
                continue
        if not already:
            do_merge = False
            val = None
            if h1 and h2 and str(h1).strip() == str(h2).strip():
                do_merge = True
                val = str(h1).strip()
            elif h1 and not h2:
                do_merge = True
                val = str(h1).strip()
            elif not h1 and h2:
                do_merge = True
                val = str(h2).strip()
            if do_merge and val:
                try:
                    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
                    top = ws.cell(row=1, column=col)
                    top.value = val
                    top.alignment = Alignment(horizontal="center", vertical="center")
                except Exception:
                    pass
        col += 2
    try:
        autosize_date_column(ws)
    except Exception:
        try:
            ws.column_dimensions["A"].width = 12
        except Exception:
            pass
    wb.save(EXCEL_FILE)
    return wb


def autosize_date_column(ws):
    max_len = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v:
            max_len = max(max_len, len(str(v)))
    ws.column_dimensions["A"].width = max(max_len + 2, 12)


def find_employee(ws, name):
    name_norm = str(name).strip()
    for c in range(2, ws.max_column + 1):
        hdr = ws.cell(row=1, column=c).value
        if hdr is None:
            continue
        if str(hdr).strip() == name_norm:
            return c, c + 1
    return None, None


def add_employee(ws, name):
    last = ws.max_column
    k = last + 1
    d = last + 2
    ws.cell(row=2, column=k, value="Keldim")
    ws.cell(row=2, column=d, value="Ketdim")
    try:
        ws.merge_cells(start_row=1, start_column=k, end_row=1, end_column=d)
        top = ws.cell(row=1, column=k)
        top.value = name
        top.alignment = Alignment(horizontal="center", vertical="center")
    except Exception:
        ws.cell(row=1, column=k, value=name)
        ws.cell(row=1, column=k).alignment = Alignment(horizontal="center", vertical="center")
    return k, d


def get_date_row(ws, date_str):
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        if str(v).strip() == date_str:
            return r
    new_r = max(ws.max_row + 1, 3)
    ws.cell(row=new_r, column=1, value=date_str)
    return new_r


def unmerge_overlapping(ws, row, c1, c2):
    for m in list(ws.merged_cells.ranges):
        try:
            if m.min_row <= row <= m.max_row and not (m.max_col < c1 or m.min_col > c2):
                ws.unmerge_cells(str(m))
        except Exception:
            continue


def write_attendance(name, action):
    now = datetime.now(TZ)
    date_str = now.strftime(DATE_FORMAT)
    time_str = now.strftime(TIME_FORMAT)
    with excel_lock:
        wb = load_or_create()
        ws = wb.active
        k_col, ket_col = find_employee(ws, name)
        if k_col is None:
            k_col, ket_col = add_employee(ws, name)
        row = get_date_row(ws, date_str)
        k_val = ws.cell(row=row, column=k_col).value
        ket_val = ws.cell(row=row, column=ket_col).value
        k_is_sabbli = bool(k_val and str(k_val).strip().upper() == "SABABLI")
        ket_is_sabbli = bool(ket_val and str(ket_val).strip().upper() == "SABABLI")
        any_sabbli = k_is_sabbli or ket_is_sabbli
        if action == "Keldim":
            if any_sabbli:
                return "already_sabbli"
            if k_val is not None and str(k_val).strip() != "":
                return "already_recorded"
            ws.cell(row=row, column=k_col, value=time_str)
            autosize_date_column(ws)
            wb.save(EXCEL_FILE)
            return "ok_keldim"
        if action == "Ketdim":
            if any_sabbli:
                return "already_sabbli"
            if ket_val is not None and str(ket_val).strip() != "":
                return "already_recorded"
            ws.cell(row=row, column=ket_col, value=time_str)
            autosize_date_column(ws)
            wb.save(EXCEL_FILE)
            return "ok_ketdim"
        if action == "Sababli":
            if any_sabbli:
                return "already_sabbli"
            ws.cell(row=row, column=k_col, value=None)
            ws.cell(row=row, column=ket_col, value=None)
            unmerge_overlapping(ws, row, k_col, ket_col)
            ws.merge_cells(start_row=row, start_column=k_col, end_row=row, end_column=ket_col)
            top = ws.cell(row=row, column=k_col)
            top.value = "SABABLI"
            top.alignment = Alignment(horizontal="center", vertical="center")
            autosize_date_column(ws)
            wb.save(EXCEL_FILE)
            return "ok_sabbli"
        raise ValueError("Unknown action")


def rows_between(start_date, end_date):
    wb = load_or_create()
    ws = wb.active
    out = []
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        try:
            d = datetime.strptime(str(v).strip(), DATE_FORMAT).date()
        except Exception:
            continue
        if start_date <= d <= end_date:
            out.append(d.strftime(DATE_FORMAT))
    return out


def build_report(rows):
    wb = load_or_create()
    ws = wb.active
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Report"
    for c in range(1, ws.max_column + 1):
        new_ws.cell(row=1, column=c, value=ws.cell(row=1, column=c).value)
        new_ws.cell(row=2, column=c, value=ws.cell(row=2, column=c).value)
    dest = 3
    mapping = {}
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value in rows:
            for c in range(1, ws.max_column + 1):
                new_ws.cell(row=dest, column=c, value=ws.cell(r, c).value)
            mapping[r] = dest
            dest += 1
    try:
        for m in ws.merged_cells.ranges:
            try:
                min_row = m.min_row
                max_row = m.max_row
                min_col = m.min_col
                max_col = m.max_col
                if max_row <= 2:
                    new_ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                else:
                    if min_row == max_row and min_row in mapping:
                        new_row = mapping[min_row]
                        new_ws.merge_cells(start_row=new_row, start_column=min_col, end_row=new_row, end_column=max_col)
            except Exception:
                continue
    except Exception:
        pass
    autosize_date_column(new_ws)
    bio = BytesIO()
    new_wb.save(bio)
    bio.seek(0)
    return bio


def keyboard_for(uid):
    base = [[KeyboardButton(BTN_KELDIM), KeyboardButton(BTN_KETDIM), KeyboardButton(BTN_SABBLI)]]
    if uid in ADMINS:
        base += [[KeyboardButton(BTN_WEEK), KeyboardButton(BTN_MONTH)], [KeyboardButton(BTN_ASK)]]
    return ReplyKeyboardMarkup(base, resize_keyboard=True)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id if update.effective_user else 0
    await update.message.reply_text("Salom! Davomat botga xush kelibsiz.\nTugmalardan birini tanlang:", reply_markup=keyboard_for(uid))


async def admin_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message is None:
        return ConversationHandler.END
    raw = update.message.text or ""
    text = _normalize(raw)
    if text == _normalize(BTN_ASK):
        await update.message.reply_text("Iltimos ikkita sanani kiriting (masalan):\n2025-09-01 2025-09-15\nFormat: YYYY-MM-DD YYYY-MM-DD")
        return WAITING_DATES
    if text == _normalize(BTN_WEEK) or "HAFTALIK" in text.upper():
        end = datetime.now(TZ).date()
        start = end - timedelta(days=6)
    elif text == _normalize(BTN_MONTH) or "OYLIK" in text.upper():
        end = datetime.now(TZ).date()
        start = date(end.year, end.month, 1)
    else:
        return await handle_user(update, context)
    rows = rows_between(start, end)
    if not rows:
        await update.message.reply_text("Ma'lumot topilmadi.")
        return ConversationHandler.END
    bio = build_report(rows)
    fname = f"attendance_{start.strftime(DATE_FORMAT)}_to_{end.strftime(DATE_FORMAT)}.xlsx"
    await update.message.reply_document(document=InputFile(bio, filename=fname))
    return ConversationHandler.END


async def receive_dates(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message is None:
        return ConversationHandler.END
    parts = (update.message.text or "").split()
    if len(parts) != 2:
        await update.message.reply_text("Noto'g'ri format. Iltimos: YYYY-MM-DD YYYY-MM-DD")
        return ConversationHandler.END
    try:
        s = datetime.strptime(parts[0], DATE_FORMAT).date()
        e = datetime.strptime(parts[1], DATE_FORMAT).date()
    except Exception:
        await update.message.reply_text("Sana formatida xato. Iltimos YYYY-MM-DD formatida kiriting.")
        return ConversationHandler.END
    if s > e:
        await update.message.reply_text("Birinchi sana ikkinchi sanadan katta bo'lmasligi kerak.")
        return ConversationHandler.END
    rows = rows_between(s, e)
    if not rows:
        await update.message.reply_text("Ma'lumot topilmadi.")
        return ConversationHandler.END
    bio = build_report(rows)
    fname = f"attendance_{s.strftime(DATE_FORMAT)}_to_{e.strftime(DATE_FORMAT)}.xlsx"
    await update.message.reply_document(document=InputFile(bio, filename=fname))
    return ConversationHandler.END


async def handle_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message is None:
        return
    raw = update.message.text or ""
    text = _normalize(raw)
    user = update.effective_user
    name = user.full_name if user else "NoName"
    if text == _normalize(BTN_KELDIM):
        res = write_attendance(name, "Keldim")
    elif text == _normalize(BTN_KETDIM):
        res = write_attendance(name, "Ketdim")
    elif text == _normalize(BTN_SABBLI):
        res = write_attendance(name, "Sababli")
    else:
        if "✅" in raw:
            res = write_attendance(name, "Keldim")
        elif "🚪" in raw:
            res = write_attendance(name, "Ketdim")
        elif "⚠" in raw:
            res = write_attendance(name, "Sababli")
        else:
            await update.message.reply_text("Noma'lum buyruq. Iltimos tugmalardan foydalaning.")
            return
    if res == "already_recorded":
        await update.message.reply_text("Bugungi holatingiz allaqachon qayd etilgan.")
    elif res == "already_sabbli":
        await update.message.reply_text("Bugun siz sababli deb belgilangan.")
    elif res == "ok_keldim":
        now = datetime.now(TZ)
        await update.message.reply_text(f"✅ {name} — Keldim ({now.strftime(TIME_FORMAT)})")
    elif res == "ok_ketdim":
        now = datetime.now(TZ)
        await update.message.reply_text(f"🚪 {name} — Ketdim ({now.strftime(TIME_FORMAT)})")
    elif res == "ok_sabbli":
        await update.message.reply_text("⚠️ SABABLI belgilandi.")
    else:
        await update.message.reply_text("Xato: yozishda muammo yuz berdi.")


async def cmd_daily(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if user is None or user.id not in ADMINS:
        await update.message.reply_text("Siz admin emassiz")
        return
    today = datetime.now(TZ).date()
    rows = rows_between(today, today)
    if not rows:
        await update.message.reply_text("Bugun uchun ma'lumot yo‘q.")
        return
    bio = build_report(rows)
    fname = f"attendance_{today.strftime(DATE_FORMAT)}.xlsx"
    await update.message.reply_document(document=InputFile(bio, filename=fname))


async def cmd_week(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if user is None or user.id not in ADMINS:
        await update.message.reply_text("Siz admin emassiz")
        return
    end = datetime.now(TZ).date()
    start = end - timedelta(days=6)
    rows = rows_between(start, end)
    if not rows:
        await update.message.reply_text("Ma'lumot topilmadi.")
        return
    bio = build_report(rows)
    fname = f"attendance_{start.strftime(DATE_FORMAT)}_to_{end.strftime(DATE_FORMAT)}.xlsx"
    await update.message.reply_document(document=InputFile(bio, filename=fname))


async def cmd_month(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if user is None or user.id not in ADMINS:
        await update.message.reply_text("Siz admin emassiz")
        return
    today = datetime.now(TZ).date()
    start = date(today.year, today.month, 1)
    if today.month == 12:
        next_month = date(today.year + 1, 1, 1)
    else:
        next_month = date(today.year, today.month + 1, 1)
    end = next_month - timedelta(days=1)
    rows = rows_between(start, end)
    if not rows:
        await update.message.reply_text("Ma'lumot topilmadi.")
        return
    bio = build_report(rows)
    fname = f"attendance_{today.strftime('%Y_%m')}.xlsx"
    await update.message.reply_document(document=InputFile(bio, filename=fname))


def main():
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN required")
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("daily", cmd_daily))
    app.add_handler(CommandHandler("week", cmd_week))
    app.add_handler(CommandHandler("month", cmd_month))
    conv_filter = filters.User(user_id=list(ADMINS)) & filters.TEXT & ~filters.COMMAND
    conv = ConversationHandler(
        entry_points=[MessageHandler(conv_filter, admin_entry)],
        states={WAITING_DATES: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_dates)]},
        fallbacks=[],
        allow_reentry=True,
    )
    app.add_handler(conv)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user))
    logger.info("Bot started (EXCEL only, TZ=Asia/Tashkent)")
    app.run_polling(allowed_updates=None)


if __name__ == "__main__":
    main()
